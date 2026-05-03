"""
Excel导出服务
使用openpyxl生成专业格式Excel报告
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Optional
from app.database import get_db_connection


def generate_excel_report(company_id: int, report_type: str = "monthly") -> bytes:
    """
    生成Excel碳排报告
    
    Args:
        company_id: 企业ID
        report_type: 报告类型 monthly/quarterly/annual
    
    Returns:
        bytes: Excel文件内容
    """
    # 获取企业信息
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM companies WHERE id = ?", (company_id,))
    company = cursor.fetchone()
    if not company:
        conn.close()
        raise ValueError("企业不存在")

    # 获取碳记录
    cursor.execute("""
        SELECT * FROM carbon_records WHERE company_id = ?
        ORDER BY record_date DESC, scope
    """, (company_id,))
    records = [dict(r) for r in cursor.fetchall()]
    conn.close()

    wb = Workbook()
    _build_summary_sheet(wb, dict(company), records)
    _build_detail_sheet(wb, records)
    _build_scope_sheet(wb, records)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary_sheet(wb: Workbook, company: dict, records: list):
    """构建汇总表"""
    ws = wb.active
    ws.title = "碳排放汇总"

    # 样式定义
    title_font = Font(name='SimHei', size=16, bold=True)
    header_font = Font(name='SimHei', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    label_font = Font(name='SimHei', size=11, bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 标题
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{company.get('name', '企业')} 碳排放汇总报告"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # 企业信息
    info_data = [
        ("企业名称", company.get('name', '')),
        ("所属行业", company.get('industry', '')),
        ("所在地区", company.get('region', '')),
        ("报告生成时间", __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')),
    ]
    for i, (label, value) in enumerate(info_data, start=3):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=2, value=value)

    # 汇总统计
    scope1 = sum(r['co2_emission'] for r in records if r['scope'] == 'scope1')
    scope2 = sum(r['co2_emission'] for r in records if r['scope'] == 'scope2')
    scope3 = sum(r['co2_emission'] for r in records if r['scope'] == 'scope3')
    total = scope1 + scope2 + scope3

    row = 8
    ws.cell(row=row, column=1, value="碳排放范围").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2, value="排放量(kgCO2)").font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    ws.cell(row=row, column=3, value="占比").font = header_font
    ws.cell(row=row, column=3).fill = header_fill
    for col in range(1, 4):
        ws.cell(row=row, column=col).border = border

    summary_rows = [
        ("范围1 直接排放", scope1),
        ("范围2 间接排放(电力)", scope2),
        ("范围3 其他间接排放", scope3),
        ("合计", total),
    ]
    for i, (name, val) in enumerate(summary_rows, start=row + 1):
        ws.cell(row=i, column=1, value=name).font = label_font if i == row + 4 else Font(name='SimHei')
        ws.cell(row=i, column=2, value=round(val, 2)).number_format = '#,##0.00'
        ws.cell(row=i, column=3, value=f"{val / total * 100:.1f}%" if total > 0 else "0%")
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = border

    # 列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15


def _build_detail_sheet(wb: Workbook, records: list):
    """构建明细表"""
    ws = wb.create_sheet("排放明细")

    header_font = Font(name='SimHei', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["序号", "记录月份", "排放范围", "排放源", "消耗量", "单位", "排放因子", "碳排放量(kgCO2)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    source_map = {
        'natural_gas': '天然气', 'coal': '煤炭', 'electricity': '外购电力',
        'gasoline': '汽油', 'diesel': '柴油', 'renewable': '绿电',
        'business_flight_short': '短途航班', 'business_flight_medium': '中途航班',
        'business_flight_long': '长途航班', 'business_train': '火车',
        'business_car': '公务汽车', 'waste_landfill': '填埋处理',
        'waste_incineration': '焚烧处理', 'waste_composting': '堆肥处理',
        'purchased_office': '办公用品', 'purchased_equipment': '设备采购',
    }
    scope_map = {'scope1': '范围1', 'scope2': '范围2', 'scope3': '范围3'}

    for i, r in enumerate(records, start=2):
        ws.cell(row=i, column=1, value=i - 1).border = border
        ws.cell(row=i, column=2, value=r['record_date']).border = border
        ws.cell(row=i, column=3, value=scope_map.get(r['scope'], r['scope'])).border = border
        ws.cell(row=i, column=4, value=source_map.get(r['emission_source'], r['emission_source'])).border = border
        ws.cell(row=i, column=5, value=r['quantity']).border = border
        ws.cell(row=i, column=6, value=r['unit']).border = border
        ws.cell(row=i, column=7, value=r.get('emission_factor', '')).border = border
        ws.cell(row=i, column=8, value=round(r['co2_emission'], 2)).border = border
        ws.cell(row=i, column=8).number_format = '#,##0.00'

    # 列宽
    widths = [8, 12, 12, 15, 12, 8, 12, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_scope_sheet(wb: Workbook, records: list):
    """构建月度范围分析表"""
    ws = wb.create_sheet("月度范围分析")

    header_font = Font(name='SimHei', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["月份", "范围1(kgCO2)", "范围2(kgCO2)", "范围3(kgCO2)", "合计(kgCO2)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # 按月聚合
    monthly = {}
    for r in records:
        month = r['record_date']
        if month not in monthly:
            monthly[month] = {'scope1': 0, 'scope2': 0, 'scope3': 0}
        monthly[month][r['scope']] += r['co2_emission']

    for i, (month, data) in enumerate(sorted(monthly.items()), start=2):
        ws.cell(row=i, column=1, value=month).border = border
        ws.cell(row=i, column=2, value=round(data['scope1'], 2)).border = border
        ws.cell(row=i, column=3, value=round(data['scope2'], 2)).border = border
        ws.cell(row=i, column=4, value=round(data['scope3'], 2)).border = border
        ws.cell(row=i, column=5, value=round(data['scope1'] + data['scope2'] + data['scope3'], 2)).border = border
        for col in range(2, 6):
            ws.cell(row=i, column=col).number_format = '#,##0.00'

    widths = [12, 18, 18, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
